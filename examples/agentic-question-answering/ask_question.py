import argparse
import csv
import json
import os
import sqlite3
import tempfile
from pathlib import Path
from typing import Any

from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries


STRUCTURED_SYSTEM_PROMPT = """
You answer questions about a spreadsheet that has already been transformed into structured tables.

Always follow this workflow:
1. Call get_table_overview first.
2. Use run_sql_query to inspect the relevant tables.
3. Give a concise answer.
4. Always cite cell references from columns prefixed with __ref or __refs whenever they are available.

Never guess. If the data is insufficient, say so.
""".strip()

EXCEL_SYSTEM_PROMPT = """
You answer questions directly from a raw Excel workbook.

Always follow this workflow:
1. Call list_excel_sheets or search_excel_values first.
2. Read only the ranges you need with read_excel_range.
3. Give a concise answer.
4. Always cite cell references.

Never guess. If the data is insufficient, say so.
""".strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", choices=["excel", "sqlite", "csv"], required=True)
    parser.add_argument("--question", required=True)
    parser.add_argument("--model", default=os.environ.get("OPENAI_MODEL", "gpt-4.1"))
    parser.add_argument("--excel-path")
    parser.add_argument("--sqlite-path")
    parser.add_argument("--csv-dir")
    return parser.parse_args()


def normalize_sqlite_type(type_name: str | None) -> str:
    upper_type = (type_name or "TEXT").upper()
    if "INT" in upper_type:
        return "INTEGER"
    if any(token in upper_type for token in ["REAL", "FLOAT", "DOUBLE", "NUMERIC", "DECIMAL"]):
        return "REAL"
    return "TEXT"


def quote_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def maybe_parse_numeric(value: str, type_name: str) -> Any:
    if value == "":
        return None
    normalized_type = normalize_sqlite_type(type_name)
    if normalized_type == "INTEGER":
        return int(value)
    if normalized_type == "REAL":
        return float(value)
    return value


class StructuredSource:
    def __init__(self, *, sqlite_path: Path | None = None, csv_dir: Path | None = None) -> None:
        self._temp_dir: tempfile.TemporaryDirectory[str] | None = None
        if sqlite_path is not None:
            self.connection = sqlite3.connect(sqlite_path)
        elif csv_dir is not None:
            self._temp_dir = tempfile.TemporaryDirectory()
            temp_sqlite_path = Path(self._temp_dir.name) / "structured_outputs.sqlite"
            self.connection = sqlite3.connect(temp_sqlite_path)
            self._load_csv_dir(csv_dir)
        else:
            raise ValueError("Expected either sqlite_path or csv_dir.")

        self.connection.row_factory = sqlite3.Row

    def close(self) -> None:
        self.connection.close()
        if self._temp_dir is not None:
            self._temp_dir.cleanup()

    def get_table_overview(self) -> dict[str, Any]:
        cursor = self.connection.execute(
            "SELECT table_name, table_type, column_name, column_sqlite_type, column_name_refs "
            "FROM __deeptable_table_overview ORDER BY table_name, rowid"
        )
        tables: dict[str, dict[str, Any]] = {}
        for row in cursor.fetchall():
            table_name = row["table_name"]
            table = tables.setdefault(
                table_name,
                {"table_name": table_name, "table_type": row["table_type"], "columns": []},
            )
            table["columns"].append(
                {
                    "name": row["column_name"],
                    "sqlite_type": row["column_sqlite_type"],
                    "refs": row["column_name_refs"],
                }
            )
        return {"tables": list(tables.values())}

    def run_sql_query(self, query: str) -> dict[str, Any]:
        normalized_query = query.lstrip().upper()
        if not (normalized_query.startswith("SELECT") or normalized_query.startswith("WITH")):
            return {"error": "Only SELECT and WITH queries are allowed."}

        cursor = self.connection.execute(query)
        rows = cursor.fetchall()
        columns = [description[0] for description in cursor.description or []]
        serialized_rows = [dict(zip(columns, row, strict=False)) for row in rows[:200]]
        return {
            "columns": columns,
            "row_count": len(rows),
            "rows": serialized_rows,
            "truncated": len(rows) > 200,
        }

    def _load_csv_dir(self, csv_dir: Path) -> None:
        schema_by_table: dict[str, dict[str, str]] = {}
        overview_path = csv_dir / "__deeptable_table_overview.csv"
        if overview_path.exists():
            with overview_path.open(newline="", encoding="utf-8") as handle:
                for row in csv.DictReader(handle):
                    table_name = row["table_name"]
                    schema_by_table.setdefault(table_name, {})[row["column_name"]] = row["column_sqlite_type"]

        for csv_path in sorted(csv_dir.glob("*.csv")):
            with csv_path.open(newline="", encoding="utf-8") as handle:
                reader = csv.DictReader(handle)
                fieldnames = reader.fieldnames or []
                table_name = csv_path.stem
                declared_types = schema_by_table.get(table_name, {})
                columns_sql = ", ".join(
                    f"{quote_identifier(field)} {normalize_sqlite_type(declared_types.get(field))}"
                    for field in fieldnames
                )
                self.connection.execute(f"CREATE TABLE {quote_identifier(table_name)} ({columns_sql})")
                insert_sql = (
                    f"INSERT INTO {quote_identifier(table_name)} VALUES "
                    f"({', '.join('?' for _ in fieldnames)})"
                )
                rows_to_insert = []
                for row in reader:
                    rows_to_insert.append(
                        [
                            maybe_parse_numeric(row.get(field, ""), declared_types.get(field, "TEXT"))
                            for field in fieldnames
                        ]
                    )
                self.connection.executemany(insert_sql, rows_to_insert)
        self.connection.commit()


class ExcelSource:
    def __init__(self, excel_path: Path) -> None:
        self.workbook = load_workbook(excel_path, data_only=True, read_only=True)

    def close(self) -> None:
        self.workbook.close()

    def list_sheets(self) -> dict[str, Any]:
        sheets = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheets.append(
                {
                    "sheet_name": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                }
            )
        return {"sheets": sheets}

    def search_values(self, search_text: str, max_results: int = 25) -> dict[str, Any]:
        lowered = search_text.lower()
        matches: list[dict[str, Any]] = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    value_text = str(value)
                    if lowered in value_text.lower():
                        matches.append(
                            {
                                "sheet_name": sheet_name,
                                "cell": cell.coordinate,
                                "value": value_text,
                            }
                        )
                        if len(matches) >= max_results:
                            return {"matches": matches, "truncated": True}
        return {"matches": matches, "truncated": False}

    def read_range(self, sheet_name: str, start_cell: str, end_cell: str) -> dict[str, Any]:
        sheet = self.workbook[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        rows = []
        for row_index, row in enumerate(
            sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            ),
            start=min_row,
        ):
            rows.append(
                [
                    {
                        "cell": (
                            f"{get_column_letter(column_index)}{row_index}"
                            if isinstance(cell, EmptyCell)
                            else cell.coordinate
                        ),
                        "value": None if cell.value is None else str(cell.value),
                    }
                    for column_index, cell in enumerate(row, start=min_col)
                ]
            )
        return {
            "sheet_name": sheet_name,
            "range": f"{start_cell}:{end_cell}",
            "rows": rows,
        }


def model_dump_compat(value: Any) -> Any:
    if hasattr(value, "model_dump"):
        return value.model_dump()
    if hasattr(value, "to_dict"):
        return value.to_dict()
    return value


def build_tools(source: str) -> list[dict[str, Any]]:
    if source in {"sqlite", "csv"}:
        return [
            {
                "type": "function",
                "function": {
                    "name": "get_table_overview",
                    "description": "Return the DeepTable table overview with table names, column names, and column types.",
                    "parameters": {"type": "object", "properties": {}, "required": []},
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "run_sql_query",
                    "description": "Run a read-only SQL query against the structured tables.",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "query": {"type": "string"},
                        },
                        "required": ["query"],
                    },
                },
            },
        ]

    return [
        {
            "type": "function",
            "function": {
                "name": "list_excel_sheets",
                "description": "List the workbook sheets and their approximate size.",
                "parameters": {"type": "object", "properties": {}, "required": []},
            },
        },
        {
            "type": "function",
            "function": {
                "name": "search_excel_values",
                "description": "Search the workbook for cell values containing a given text fragment.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "search_text": {"type": "string"},
                        "max_results": {"type": "integer", "default": 25},
                    },
                    "required": ["search_text"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_excel_range",
                "description": "Read a rectangular cell range from a workbook sheet.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "sheet_name": {"type": "string"},
                        "start_cell": {"type": "string"},
                        "end_cell": {"type": "string"},
                    },
                    "required": ["sheet_name", "start_cell", "end_cell"],
                },
            },
        },
    ]


def handle_tool_call(source_impl: StructuredSource | ExcelSource, tool_name: str, arguments_json: str) -> dict[str, Any]:
    arguments = json.loads(arguments_json or "{}")
    if tool_name == "get_table_overview" and isinstance(source_impl, StructuredSource):
        return source_impl.get_table_overview()
    if tool_name == "run_sql_query" and isinstance(source_impl, StructuredSource):
        return source_impl.run_sql_query(arguments["query"])
    if tool_name == "list_excel_sheets" and isinstance(source_impl, ExcelSource):
        return source_impl.list_sheets()
    if tool_name == "search_excel_values" and isinstance(source_impl, ExcelSource):
        return source_impl.search_values(
            arguments["search_text"],
            arguments.get("max_results", 25),
        )
    if tool_name == "read_excel_range" and isinstance(source_impl, ExcelSource):
        return source_impl.read_range(
            arguments["sheet_name"],
            arguments["start_cell"],
            arguments["end_cell"],
        )
    return {"error": f"Unsupported tool call: {tool_name}"}


def run_agent(source: str, source_impl: StructuredSource | ExcelSource, question: str, model: str) -> str:
    client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
    tools = build_tools(source)
    system_prompt = STRUCTURED_SYSTEM_PROMPT if source in {"sqlite", "csv"} else EXCEL_SYSTEM_PROMPT
    messages: list[dict[str, Any]] = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": question},
    ]

    for _ in range(12):
        completion = client.chat.completions.create(
            model=model,
            messages=messages,
            tools=tools,
        )
        message = completion.choices[0].message
        tool_calls = list(message.tool_calls or [])

        if not tool_calls:
            return message.content or ""

        messages.append(
            {
                "role": "assistant",
                "content": message.content or "",
                "tool_calls": [model_dump_compat(tool_call) for tool_call in tool_calls],
            }
        )

        for tool_call in tool_calls:
            result = handle_tool_call(source_impl, tool_call.function.name, tool_call.function.arguments)
            messages.append(
                {
                    "role": "tool",
                    "tool_call_id": tool_call.id,
                    "content": json.dumps(result, ensure_ascii=True),
                }
            )

    raise RuntimeError("The model did not finish within 12 tool iterations.")


def build_source(args: argparse.Namespace) -> StructuredSource | ExcelSource:
    if args.source == "excel":
        if not args.excel_path:
            raise ValueError("--excel-path is required when --source excel is used.")
        return ExcelSource(Path(args.excel_path).expanduser().resolve())
    if args.source == "sqlite":
        if not args.sqlite_path:
            raise ValueError("--sqlite-path is required when --source sqlite is used.")
        return StructuredSource(sqlite_path=Path(args.sqlite_path).expanduser().resolve())
    if not args.csv_dir:
        raise ValueError("--csv-dir is required when --source csv is used.")
    return StructuredSource(csv_dir=Path(args.csv_dir).expanduser().resolve())


def main() -> None:
    args = parse_args()
    source_impl = build_source(args)
    try:
        answer = run_agent(args.source, source_impl, args.question, args.model)
    finally:
        source_impl.close()
    print(answer)


if __name__ == "__main__":
    main()